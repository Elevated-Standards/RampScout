import os
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import re

def write_to_excel(template_path, data_frames, column_mapping, output_dir, cloud_provider="aws"):
    """
    Write processed data to an Excel template and save the file.

    :param template_path: Path to the Excel template.
    :param data_frames: List of DataFrames to consolidate and write.
    :param column_mapping: Mapping of DataFrame columns to Excel columns.
    :param output_dir: Directory to save the output file.
    """
    # Load the Excel template
    wb = load_workbook(template_path)
    sheet = wb["Inventory"]

    # Combine all DataFrames
    final_df = pd.concat(data_frames, ignore_index=True)
    final_df = final_df.rename(columns={v: k for k, v in column_mapping.items()})

    # Ensure "UNIQUE ASSET IDENTIFIER" column is filled
    if "UNIQUE ASSET IDENTIFIER" not in final_df.columns:
        final_df["UNIQUE ASSET IDENTIFIER"] = final_df["unique_id"]

    # Write data to the sheet
    for index, row in final_df.iterrows():
        for col, value in row.items():
            col_idx = list(column_mapping.keys()).index(col) + 1  # Excel is 1-indexed
            sheet.cell(row=index + 6, column=col_idx).value = value

    # Post-process to move contents from column A to column B for non-EC2 and non-RDS services (AWS specific)
    if cloud_provider == "aws":
        pattern = re.compile(r"cmd-production\d+-app\d+-\d+")
        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=1, max_col=1):
            cell_value = row[0].value
            if cell_value and not cell_value.startswith("i-") and not cell_value.startswith("cmd-"):
                sheet.cell(row=row[0].row, column=2).value = cell_value
                row[0].value = None
            elif cell_value and pattern.match(cell_value):
                sheet.cell(row=row[0].row, column=2).value = cell_value
                row[0].value = None

    # Delete column A in the "Inventory" sheet
    sheet.delete_cols(1)


    # Save the workbook
    save_workbook(wb, output_dir)

def save_workbook(wb, output_dir):
    """
    Save the workbook to the output directory with a timestamped filename.

    :param wb: The openpyxl workbook object.
    :param output_dir: Directory to save the file.
    """
    # Generate a dynamic save path
    current_date = datetime.now()
    save_path = os.path.join(output_dir, f"{current_date.year}-{current_date.month:02}-{current_date.day:02}-Inventory.xlsx")

    # Ensure the directory exists
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    # Save the file
    wb.save(save_path)
    print(f"Excel file saved to: {save_path}")